"""Reports + exports (Excel/CSV/PDF) + bulk import."""
import io
import csv
from typing import Optional, List

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from database import db
from auth import require_perm
from utils import uid, now_iso, write_audit, denormalize_products, compute_status

router = APIRouter(tags=["reports"])


async def _fetch_products(filters: dict) -> List[dict]:
    q: dict = {}
    if filters.get("category_id"):
        q["category_id"] = filters["category_id"]
    if filters.get("location_id"):
        q["location_id"] = filters["location_id"]
    if filters.get("is_active") is not None:
        q["is_active"] = filters["is_active"]
    docs = await db.products.find(q, {"_id": 0}).sort("name", 1).to_list(5000)
    docs = await denormalize_products(docs)
    if filters.get("status"):
        docs = [d for d in docs if d["status"] == filters["status"]]
    return docs


@router.get("/reports/inventory")
async def report_inventory(
    user: dict = Depends(require_perm("report:read")),
    category_id: Optional[str] = None,
    location_id: Optional[str] = None,
    status: Optional[str] = None,
    is_active: Optional[bool] = True,
):
    filters = {"category_id": category_id, "location_id": location_id, "status": status, "is_active": is_active}
    docs = await _fetch_products(filters)
    return docs


def _stream(buf: io.BytesIO, filename: str, media: str) -> StreamingResponse:
    buf.seek(0)
    return StreamingResponse(buf, media_type=media, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/reports/inventory/export")
async def export_inventory(
    user: dict = Depends(require_perm("report:read")),
    format: str = Query("xlsx", pattern="^(xlsx|csv|pdf)$"),
    category_id: Optional[str] = None,
    location_id: Optional[str] = None,
    status: Optional[str] = None,
    is_active: Optional[bool] = True,
):
    filters = {"category_id": category_id, "location_id": location_id, "status": status, "is_active": is_active}
    docs = await _fetch_products(filters)

    rows = [{
        "SKU": d["sku"],
        "Name": d["name"],
        "Category": d.get("category_name") or "",
        "Brand": d.get("brand_name") or "",
        "Location": d.get("location_name") or "",
        "Supplier": d.get("supplier_name") or "",
        "Stock": d.get("stock", 0),
        "Min": d.get("min_stock", 0),
        "Max": d.get("max_stock", 0),
        "Unit": d.get("unit", ""),
        "Unit Cost": d.get("unit_cost", 0),
        "Status": d.get("status", ""),
    } for d in docs]

    if format == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()) if rows else ["SKU"])
        w.writeheader()
        for r in rows:
            w.writerow(r)
        out = io.BytesIO(buf.getvalue().encode())
        await write_audit(user, "report.export", "report", None, {"kind": "inventory", "format": "csv", "count": len(rows)})
        return _stream(out, "inventory.csv", "text/csv")

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        headers = list(rows[0].keys()) if rows else ["SKU", "Name", "Stock"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0F5B60")
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left")
        for row in rows:
            ws.append(list(row.values()))
        for col_idx, h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(12, min(28, len(h) + 4))
        buf = io.BytesIO()
        wb.save(buf)
        await write_audit(user, "report.export", "report", None, {"kind": "inventory", "format": "xlsx", "count": len(rows)})
        return _stream(buf, "inventory.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # PDF
    buf = io.BytesIO()
    doc_ = SimpleDocTemplate(buf, pagesize=landscape(LETTER), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [Paragraph("IT Inventory Report", styles["Title"]), Spacer(1, 8)]
    table_data = [list(rows[0].keys())] if rows else [["No data"]]
    for r in rows:
        table_data.append([str(v) for v in r.values()])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B60")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ]))
    elements.append(t)
    doc_.build(elements)
    await write_audit(user, "report.export", "report", None, {"kind": "inventory", "format": "pdf", "count": len(rows)})
    return _stream(buf, "inventory.pdf", "application/pdf")


@router.get("/reports/movements/export")
async def export_movements(
    user: dict = Depends(require_perm("report:read")),
    format: str = Query("xlsx", pattern="^(xlsx|csv|pdf)$"),
    type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    q: dict = {}
    if type:
        q["type"] = type
    if date_from or date_to:
        q["created_at"] = {}
        if date_from:
            q["created_at"]["$gte"] = date_from
        if date_to:
            q["created_at"]["$lte"] = date_to
    docs = await db.inventory_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    rows = [{
        "Date": d.get("created_at", ""),
        "Type": d.get("type", ""),
        "Product": d.get("product_name", ""),
        "SKU": d.get("product_sku", ""),
        "Qty": d.get("qty", 0),
        "Previous": d.get("previous_stock", 0),
        "Resulting": d.get("resulting_stock", 0),
        "User": d.get("user_name", ""),
        "Reason": d.get("reason", ""),
        "Reference": d.get("reference", ""),
    } for d in docs]

    if format == "csv":
        buf = io.StringIO()
        if rows:
            w = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
            w.writeheader()
            for r in rows:
                w.writerow(r)
        out = io.BytesIO(buf.getvalue().encode())
        return _stream(out, "movements.csv", "text/csv")

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Movements"
        if rows:
            headers = list(rows[0].keys())
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0F5B60")
            for r in rows:
                ws.append(list(r.values()))
        buf = io.BytesIO()
        wb.save(buf)
        return _stream(buf, "movements.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # PDF
    buf = io.BytesIO()
    doc_ = SimpleDocTemplate(buf, pagesize=landscape(LETTER), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Movement History", styles["Title"]), Spacer(1, 8)]
    table_data = [list(rows[0].keys())] if rows else [["No data"]]
    for r in rows:
        table_data.append([str(v) for v in r.values()])
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F5B60")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    elements.append(t)
    doc_.build(elements)
    return _stream(buf, "movements.pdf", "application/pdf")


# ---------- Import ----------
IMPORT_COLUMNS = ["sku", "name", "description", "category", "brand", "location", "supplier", "model", "part_number", "unit", "min_stock", "max_stock", "unit_cost", "initial_stock", "barcode"]


@router.get("/import/template")
async def import_template(user: dict = Depends(require_perm("import:write"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    for i, h in enumerate(IMPORT_COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F5B60")
    # Example row
    ws.append(["KB-001", "USB Keyboard", "Standard wired keyboard", "Peripherals", "Logitech", "IT Storage", "TechSupplier Co.", "K120", "920-002478", "unit", 5, 100, 12.99, 30, "KB-001"])
    buf = io.BytesIO()
    wb.save(buf)
    return _stream(buf, "products_template.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


async def _resolve_by_name(collection: str, name: str) -> Optional[str]:
    if not name:
        return None
    doc = await db[collection].find_one({"name": name}, {"_id": 0, "id": 1})
    return doc["id"] if doc else None


@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...), user: dict = Depends(require_perm("import:write"))):
    content = await file.read()
    try:
        if file.filename.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(400, detail=f"Cannot read file: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    missing_cols = [c for c in ["sku", "name", "initial_stock"] if c not in df.columns]
    if missing_cols:
        raise HTTPException(400, detail=f"Missing required columns: {', '.join(missing_cols)}")

    valid_rows = []
    invalid_rows = []
    seen_skus = set()
    for idx, row in df.iterrows():
        errors = []
        row_dict = {c: (None if pd.isna(row[c]) else row[c]) for c in df.columns}
        sku = str(row_dict.get("sku") or "").strip()
        name = str(row_dict.get("name") or "").strip()
        if not sku:
            errors.append("sku is required")
        if not name:
            errors.append("name is required")
        if sku and sku in seen_skus:
            errors.append("duplicate sku within file")
        if sku:
            seen_skus.add(sku)
        if sku and await db.products.find_one({"sku": sku}):
            errors.append("sku already exists in database")
        try:
            initial_stock = int(row_dict.get("initial_stock") or 0)
            if initial_stock < 0:
                errors.append("initial_stock must be >= 0")
        except Exception:
            errors.append("initial_stock must be integer")
            initial_stock = 0
        try:
            min_stock = int(row_dict.get("min_stock") or 0)
        except Exception:
            min_stock = 0
        try:
            max_stock = int(row_dict.get("max_stock") or 1000)
        except Exception:
            max_stock = 1000
        try:
            unit_cost = float(row_dict.get("unit_cost") or 0.0)
        except Exception:
            unit_cost = 0.0

        parsed = {
            "sku": sku,
            "name": name,
            "description": row_dict.get("description") or None,
            "category": row_dict.get("category") or None,
            "brand": row_dict.get("brand") or None,
            "location": row_dict.get("location") or None,
            "supplier": row_dict.get("supplier") or None,
            "model": row_dict.get("model") or None,
            "part_number": row_dict.get("part_number") or None,
            "unit": row_dict.get("unit") or "unit",
            "min_stock": min_stock,
            "max_stock": max_stock,
            "unit_cost": unit_cost,
            "initial_stock": initial_stock,
            "barcode": row_dict.get("barcode") or None,
        }
        entry = {"row": int(idx) + 2, "data": parsed, "errors": errors}
        if errors:
            invalid_rows.append(entry)
        else:
            valid_rows.append(entry)
    return {"total": len(df), "valid": len(valid_rows), "invalid": len(invalid_rows), "valid_rows": valid_rows, "invalid_rows": invalid_rows}


@router.post("/import/commit")
async def import_commit(payload: dict, user: dict = Depends(require_perm("import:write"))):
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(400, detail="No rows to import.")
    created = 0
    skipped = 0
    errors: List[dict] = []
    for entry in rows:
        try:
            d = entry["data"]
            sku = d["sku"]
            if await db.products.find_one({"sku": sku}):
                skipped += 1
                continue
            doc = {
                "id": uid(),
                "sku": sku,
                "name": d["name"],
                "description": d.get("description"),
                "model": d.get("model"),
                "part_number": d.get("part_number"),
                "unit": d.get("unit") or "unit",
                "min_stock": int(d.get("min_stock") or 0),
                "max_stock": int(d.get("max_stock") or 1000),
                "unit_cost": float(d.get("unit_cost") or 0.0),
                "stock": int(d.get("initial_stock") or 0),
                "is_active": True,
                "category_id": await _resolve_by_name("categories", d.get("category") or ""),
                "brand_id": await _resolve_by_name("brands", d.get("brand") or ""),
                "location_id": await _resolve_by_name("locations", d.get("location") or ""),
                "supplier_id": await _resolve_by_name("suppliers", d.get("supplier") or ""),
                "barcode": (d.get("barcode") or sku),
                "qr_code": None,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
            doc["qr_code"] = f"QR-{doc['id']}"
            await db.products.insert_one(doc)
            if doc["stock"] > 0:
                await db.inventory_movements.insert_one({
                    "id": uid(),
                    "type": "entry",
                    "product_id": doc["id"],
                    "product_sku": doc["sku"],
                    "product_name": doc["name"],
                    "qty": doc["stock"],
                    "previous_stock": 0,
                    "resulting_stock": doc["stock"],
                    "user_id": user["id"],
                    "user_name": user.get("full_name"),
                    "reason": "bulk import",
                    "created_at": now_iso(),
                })
            created += 1
        except Exception as e:
            errors.append({"row": entry.get("row"), "error": str(e)})
    await write_audit(user, "import.commit", "import", None, {"created": created, "skipped": skipped, "errors": len(errors)})
    return {"created": created, "skipped": skipped, "errors": errors}
